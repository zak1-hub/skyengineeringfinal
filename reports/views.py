"""
reports/views.py
Student 5 – Reports (PDF / Excel)
Generates downloadable PDF and Excel reports from the organisation models.
Also provides an HTML dashboard preview.
"""
import io
from datetime import datetime

from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db.models import Count, Q
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone

from organisation.models import Department, Team


# ─────────────────────────────────────────────
# Helper: gather all report data in one place
# ─────────────────────────────────────────────
def _collect_report_data():
    departments = Department.objects.annotate(team_count=Count('team')).order_by('name')
    teams = Team.objects.select_related('department', 'manager', 'team_type').order_by('department__name', 'name')
    teams_without_manager = teams.filter(Q(manager__isnull=True))

    total_teams = teams.count()
    total_departments = departments.count()
    total_users = User.objects.count()

    # Teams per department list
    dept_summary = [
        {
            'name': d.name,
            'specialisation': d.specialisation,
            'leader': d.leader.get_full_name() if d.leader else 'N/A',
            'team_count': d.team_count,
        }
        for d in departments
    ]

    # Full team list
    team_list = [
        {
            'name': t.name,
            'department': t.department.name,
            'manager': t.manager.get_full_name() if t.manager else '—',
            'team_type': t.team_type.name if t.team_type else '—',
            'mission': t.mission or '—',
        }
        for t in teams
    ]

    # Teams without managers
    no_manager_list = [
        {
            'name': t.name,
            'department': t.department.name,
            'team_type': t.team_type.name if t.team_type else '—',
        }
        for t in teams_without_manager
    ]

    generated_at = timezone.now().strftime('%d %B %Y, %H:%M UTC')

    return {
        'total_teams': total_teams,
        'total_departments': total_departments,
        'total_users': total_users,
        'dept_summary': dept_summary,
        'team_list': team_list,
        'no_manager_list': no_manager_list,
        'no_manager_count': len(no_manager_list),
        'generated_at': generated_at,
    }


# ─────────────────────────────────────────────
# HTML Report Dashboard
# ─────────────────────────────────────────────
@login_required
def reports_home(request):
    data = _collect_report_data()
    return render(request, 'reports/reports_home.html', data)


# ─────────────────────────────────────────────
# PDF Export
# ─────────────────────────────────────────────
@login_required
def export_pdf(request):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            HRFlowable, Paragraph, SimpleDocTemplate, Spacer, Table,
            TableStyle,
        )
    except ImportError:
        return HttpResponse(
            'reportlab is not installed. Run: pip install reportlab',
            status=500,
        )

    data = _collect_report_data()
    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2 * cm,
        leftMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
        title='Sky Engineering – Team Report',
    )

    styles = getSampleStyleSheet()
    sky_dark = colors.HexColor('#003057')
    sky_blue = colors.HexColor('#00b5e2')

    title_style = ParagraphStyle(
        'SkyTitle',
        parent=styles['Title'],
        textColor=sky_dark,
        fontSize=20,
        spaceAfter=6,
    )
    h2_style = ParagraphStyle(
        'SkyH2',
        parent=styles['Heading2'],
        textColor=sky_dark,
        fontSize=13,
        spaceBefore=14,
        spaceAfter=4,
    )
    normal = styles['Normal']
    normal.fontSize = 9

    story = []

    # ── Title block ──────────────────────────────────
    story.append(Paragraph('Sky Engineering', title_style))
    story.append(Paragraph('Team & Department Report', styles['Heading2']))
    story.append(Paragraph(f'Generated: {data["generated_at"]}', normal))
    story.append(HRFlowable(width='100%', thickness=2, color=sky_blue, spaceAfter=12))

    # ── KPI summary ──────────────────────────────────
    story.append(Paragraph('Summary', h2_style))
    kpi_data = [
        ['Metric', 'Value'],
        ['Total Teams', str(data['total_teams'])],
        ['Total Departments', str(data['total_departments'])],
        ['Total Users', str(data['total_users'])],
        ['Teams Without a Manager', str(data['no_manager_count'])],
    ]
    kpi_table = Table(kpi_data, colWidths=[10 * cm, 5 * cm])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), sky_dark),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#e8f7fc')]),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#c0d8e4')),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 12))

    # ── Teams per Department ─────────────────────────
    story.append(Paragraph('Teams per Department', h2_style))
    dept_table_data = [['Department', 'Specialisation', 'Leader', '# Teams']]
    for d in data['dept_summary']:
        dept_table_data.append([d['name'], d['specialisation'], d['leader'], str(d['team_count'])])
    dept_table = Table(dept_table_data, colWidths=[5 * cm, 5 * cm, 4 * cm, 2.5 * cm])
    dept_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), sky_dark),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#e8f7fc')]),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#c0d8e4')),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(dept_table)
    story.append(Spacer(1, 12))

    # ── Full Team List ───────────────────────────────
    story.append(Paragraph('All Teams', h2_style))
    team_table_data = [['Team Name', 'Department', 'Manager', 'Type']]
    for t in data['team_list']:
        team_table_data.append([t['name'], t['department'], t['manager'], t['team_type']])
    team_table = Table(team_table_data, colWidths=[5 * cm, 4.5 * cm, 4 * cm, 3 * cm])
    team_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), sky_dark),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#e8f7fc')]),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#c0d8e4')),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(team_table)
    story.append(Spacer(1, 12))

    # ── Teams Without Managers ───────────────────────
    story.append(Paragraph('Teams Without a Manager', h2_style))
    if data['no_manager_list']:
        nm_data = [['Team Name', 'Department', 'Type']]
        for t in data['no_manager_list']:
            nm_data.append([t['name'], t['department'], t['team_type']])
        nm_table = Table(nm_data, colWidths=[6 * cm, 6 * cm, 4.5 * cm])
        nm_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#b71c1c')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fde8e8')]),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#f5c6c6')),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        story.append(nm_table)
    else:
        story.append(Paragraph('✓ All teams have an assigned manager.', normal))

    doc.build(story)
    buffer.seek(0)
    filename = f'sky_engineering_report_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf'
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─────────────────────────────────────────────
# Excel Export
# ─────────────────────────────────────────────
@login_required
def export_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse(
            'openpyxl is not installed. Run: pip install openpyxl',
            status=500,
        )

    data = _collect_report_data()

    wb = openpyxl.Workbook()

    sky_dark_hex = '003057'
    sky_blue_hex = '00b5e2'
    sky_light_hex = 'e8f7fc'
    red_hex = 'b71c1c'
    red_light_hex = 'fde8e8'

    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill_dark = PatternFill('solid', fgColor=sky_dark_hex)
    header_fill_red = PatternFill('solid', fgColor=red_hex)
    alt_fill = PatternFill('solid', fgColor=sky_light_hex)
    alt_fill_red = PatternFill('solid', fgColor=red_light_hex)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    def style_header_row(ws, row_num, fill):
        for cell in ws[row_num]:
            cell.font = header_font
            cell.fill = fill
            cell.alignment = center

    def auto_width(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 45)

    # ── Sheet 1: Summary ─────────────────────────────
    ws1 = wb.active
    ws1.title = 'Summary'
    ws1.append(['Sky Engineering – Report', ''])
    ws1.append([f'Generated: {data["generated_at"]}', ''])
    ws1.append([])
    ws1.append(['Metric', 'Value'])
    style_header_row(ws1, 4, header_fill_dark)
    summary_rows = [
        ('Total Teams', data['total_teams']),
        ('Total Departments', data['total_departments']),
        ('Total Users', data['total_users']),
        ('Teams Without a Manager', data['no_manager_count']),
    ]
    for i, (metric, value) in enumerate(summary_rows):
        ws1.append([metric, value])
        if i % 2 == 1:
            for cell in ws1[ws1.max_row]:
                cell.fill = alt_fill
    ws1['A1'].font = Font(bold=True, size=14, color=sky_dark_hex)
    auto_width(ws1)

    # ── Sheet 2: Departments ─────────────────────────
    ws2 = wb.create_sheet('Departments')
    ws2.append(['Department', 'Specialisation', 'Leader', 'Number of Teams'])
    style_header_row(ws2, 1, header_fill_dark)
    for i, d in enumerate(data['dept_summary']):
        ws2.append([d['name'], d['specialisation'], d['leader'], d['team_count']])
        if i % 2 == 1:
            for cell in ws2[ws2.max_row]:
                cell.fill = alt_fill
    auto_width(ws2)

    # ── Sheet 3: All Teams ───────────────────────────
    ws3 = wb.create_sheet('All Teams')
    ws3.append(['Team Name', 'Department', 'Manager', 'Team Type', 'Mission'])
    style_header_row(ws3, 1, header_fill_dark)
    for i, t in enumerate(data['team_list']):
        ws3.append([t['name'], t['department'], t['manager'], t['team_type'], t['mission']])
        if i % 2 == 1:
            for cell in ws3[ws3.max_row]:
                cell.fill = alt_fill
    auto_width(ws3)

    # ── Sheet 4: Teams Without Managers ─────────────
    ws4 = wb.create_sheet('No Manager')
    ws4.append(['Team Name', 'Department', 'Team Type'])
    style_header_row(ws4, 1, header_fill_red)
    for i, t in enumerate(data['no_manager_list']):
        ws4.append([t['name'], t['department'], t['team_type']])
        if i % 2 == 1:
            for cell in ws4[ws4.max_row]:
                cell.fill = alt_fill_red
    if not data['no_manager_list']:
        ws4.append(['All teams have an assigned manager.', '', ''])
    auto_width(ws4)

    # ── Deliver ──────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f'sky_engineering_report_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
